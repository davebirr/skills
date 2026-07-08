#!/usr/bin/env python3
"""
Build a cost proposal workbook (.xlsx) from markdown tables.

Usage:
    python build_cost_workbook.py <markdown_file> [output_xlsx]

Input formats:
1. Standalone markdown file with YAML frontmatter + cost tables
2. A proposal response markdown file containing <!-- cost-start --> / <!-- cost-end --> markers

The script produces a 3-sheet workbook:
  - Summary: totals by phase/category
  - Detail: full line-item breakdown (Phase, Task, Role, Hours, Customer Rate/Total, Kutlu Rate/Total)
  - Rate Card: role/rate assumptions with defaults

Rates default to $225/hr (customer) and $150/hr (Kutlu) but are editable per line item.
The Summary and Detail sheets use Excel formulas referencing the Rate Card where possible.
"""

import argparse
import os
import re
import sys

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML required. Install: pip install pyyaml")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.workbook.properties import CalcProperties
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


# ── Constants ──────────────────────────────────────────────────────

DEFAULT_CUSTOMER_RATE = 225.0
DEFAULT_KUTLU_RATE = 150.0
_UNSAFE_CHARS = re.compile(r'[<>:"/\\|?*]')


def sanitize_filename(name):
    return _UNSAFE_CHARS.sub('', name).strip()


# ── Styles ─────────────────────────────────────────────────────────

class Styles:
    TITLE = Font(name="Segoe UI", size=14, bold=True)
    SUBTITLE = Font(name="Segoe UI", size=11, color="555555")
    SECTION = Font(name="Segoe UI", size=12, bold=True)
    HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    CELL = Font(name="Segoe UI", size=10)
    BOLD_CELL = Font(name="Segoe UI", size=10, bold=True)
    INPUT = Font(name="Segoe UI", size=10, color="0000FF")  # Blue = editable input
    NOTE = Font(name="Segoe UI", size=9, italic=True, color="666666")
    CURRENCY_FMT = '$#,##0.00'
    RATE_FMT = '$#,##0.00'
    HOURS_FMT = '#,##0.0'
    THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    TOTAL_BORDER = Border(
        top=Side(style="double", color="000000"),
        bottom=Side(style="thin", color="D0D0D0"),
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
    )
    INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # Light yellow


# ── Parsing ────────────────────────────────────────────────────────

def extract_cost_markdown(text):
    """Extract cost content from <!-- cost-start --> / <!-- cost-end --> markers.
    If no markers found, treat the entire body as cost content."""
    pattern = re.compile(
        r"<!-- cost-start -->\s*\n(.*?)\n\s*<!-- cost-end -->",
        re.DOTALL,
    )
    matches = pattern.findall(text)
    if matches:
        return "\n\n".join(m.strip() for m in matches)
    return None


def parse_frontmatter(text):
    """Extract YAML frontmatter from markdown text."""
    m = re.match(r"^---\s*\n(.*?)\n---\s*\n", text, re.DOTALL)
    if m:
        data = yaml.safe_load(m.group(1)) or {}
        body = text[m.end():]
        return data, body
    return {}, text


def flatten_dict(d, prefix=""):
    """Flatten nested dict with dotted keys."""
    out = {}
    for k, v in d.items():
        key = f"{prefix}{k}" if prefix else k
        if isinstance(v, dict):
            out.update(flatten_dict(v, f"{key}."))
        else:
            out[key] = v
    return out


def parse_md_tables(text):
    """Parse markdown into list of (title, headers, rows) tuples.
    rows contain raw cell text (bold markers stripped at display time)."""
    lines = text.strip().split("\n")
    tables = []
    current_title = ""
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("## ") or line.startswith("### "):
            current_title = re.sub(r"^#+\s*", "", line).strip()
            i += 1
            continue
        if line.strip().startswith("|"):
            header_cells = [c.strip() for c in line.strip().strip("|").split("|")]
            i += 1
            if i < len(lines) and re.match(r"^\|[\s\-:|]+\|$", lines[i].strip()):
                i += 1
            rows = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                if re.match(r"^\|[\s\-:|]+\|$", lines[i].strip()):
                    i += 1
                    continue
                cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                rows.append(cells)
                i += 1
            tables.append((current_title, header_cells, rows))
            continue
        i += 1
    return tables


def classify_columns(headers):
    """Map header names to semantic roles.
    Returns dict: role -> column_index (0-based)."""
    mapping = {}
    lower_headers = [h.lower().replace("**", "").strip() for h in headers]
    for i, h in enumerate(lower_headers):
        if "phase" in h or "category" in h:
            mapping["phase"] = i
        elif "task" in h or "description" in h or "service" in h or "item" in h or "deliverable" in h:
            mapping["task"] = i
        elif "role" in h or "resource" in h or "position" in h:
            mapping["role"] = i
        elif "hour" in h:
            mapping["hours"] = i
        elif "customer" in h and ("rate" in h or "price" in h or "/hr" in h):
            mapping["customer_rate"] = i
        elif "customer" in h and ("total" in h or "cost" in h or "amount" in h):
            mapping["customer_total"] = i
        elif ("kutlu" in h or "our" in h or "internal" in h or "cost" in h) and ("rate" in h or "/hr" in h):
            mapping["kutlu_rate"] = i
        elif ("kutlu" in h or "our" in h or "internal" in h) and ("total" in h or "cost" in h or "amount" in h):
            mapping["kutlu_total"] = i
        elif "rate" in h and "customer" not in h and "kutlu" not in h:
            # Generic rate — assume customer rate
            if "customer_rate" not in mapping:
                mapping["customer_rate"] = i
        elif "total" in h or "amount" in h or "cost" in h:
            if "customer_total" not in mapping:
                mapping["customer_total"] = i
    # If no task column found, use first column
    if "task" not in mapping and 0 not in mapping.values():
        mapping["task"] = 0
    return mapping


def parse_numeric(text):
    """Try to parse a cell as a number, stripping currency and bold markers."""
    clean = text.replace("**", "").replace(",", "").replace("$", "").replace("~", "").strip()
    if not clean or clean == "-":
        return None
    try:
        return float(clean) if "." in clean else int(clean)
    except (ValueError, TypeError):
        return None


def clean_text(text):
    """Strip bold markers and whitespace."""
    return text.replace("**", "").strip()


# ── Line item structure ────────────────────────────────────────────

class LineItem:
    def __init__(self, phase="", task="", role="", hours=0,
                 customer_rate=None, kutlu_rate=None):
        self.phase = phase
        self.task = task
        self.role = role
        self.hours = hours or 0
        self.customer_rate = customer_rate
        self.kutlu_rate = kutlu_rate
        self.is_total = False

    def effective_customer_rate(self):
        return self.customer_rate if self.customer_rate is not None else DEFAULT_CUSTOMER_RATE

    def effective_kutlu_rate(self):
        return self.kutlu_rate if self.kutlu_rate is not None else DEFAULT_KUTLU_RATE


def extract_from_yaml_pricing(meta):
    """Extract LineItem objects from YAML frontmatter pricing block.

    Expected structure:
        pricing:
          rates:
            customer_facing_rate: 225
            subcontractor_bill_rate: 150
          phases:
            - id: phase_id
              name: "Phase name"
              hours: 60
              cost: 13500        (optional — ignored; we compute from rate * hours)
              role: "Consultant" (optional — defaults to "General")
          travel:               (optional — added as a separate line item)
            description: "..."
            amount: 5000
          tax:                  (optional — added as a note, not a billable line)
            rate: 0.06
            estimated_tax: 1080

    Returns list of LineItem or empty list if no pricing block found.
    """
    pricing = meta.get("pricing")
    if not pricing:
        return []

    rates = pricing.get("rates", {})
    customer_rate = rates.get("customer_facing_rate", DEFAULT_CUSTOMER_RATE)
    kutlu_rate = rates.get("subcontractor_bill_rate", DEFAULT_KUTLU_RATE)

    phases = pricing.get("phases", [])
    if not phases:
        return []

    items = []
    for phase in phases:
        name = phase.get("name", phase.get("id", ""))
        hours = phase.get("hours", 0) or 0
        role = phase.get("role", "General")
        phase_cust_rate = phase.get("customer_rate", customer_rate)
        phase_kutlu_rate = phase.get("kutlu_rate", kutlu_rate)

        # Split "Phase N — Description" into phase/task if pattern matches
        # Use em-dash/en-dash with surrounding spaces to avoid splitting "Hyper-V"
        m = re.match(r"^(Phase\s+\d+)\s*[—–]\s*(.+)$", name)
        if m:
            phase_name = m.group(1).strip()
            task_name = m.group(2).strip()
        else:
            # Try splitting on " — " or " – " (spaced dashes only)
            parts = re.split(r"\s+[—–]\s+", name, maxsplit=1)
            if len(parts) == 2:
                phase_name = parts[0].strip()
                task_name = parts[1].strip()
            else:
                phase_name = name
                task_name = name

        item = LineItem(
            phase=phase_name,
            task=task_name,
            role=role,
            hours=hours,
            customer_rate=phase_cust_rate,
            kutlu_rate=phase_kutlu_rate,
        )
        items.append(item)

    # Travel line item (if present)
    travel = pricing.get("travel")
    if travel and travel.get("amount"):
        travel_item = LineItem(
            phase="Travel",
            task=travel.get("description", "Travel expenses"),
            role="Travel",
            hours=0,
            customer_rate=0,
            kutlu_rate=0,
        )
        travel_item._fixed_customer_total = travel["amount"]
        travel_item._fixed_kutlu_total = travel["amount"]  # pass-through
        items.append(travel_item)

    return items


def extract_line_items(tables):
    """Convert parsed markdown tables into LineItem objects."""
    items = []
    current_phase = ""
    for title, headers, rows in tables:
        if title:
            current_phase = title
        col_map = classify_columns(headers)
        for row in rows:
            is_total = any("**" in cell for cell in row)
            phase = clean_text(row[col_map["phase"]]) if "phase" in col_map and col_map["phase"] < len(row) else current_phase
            task = clean_text(row[col_map["task"]]) if "task" in col_map and col_map["task"] < len(row) else ""
            role = clean_text(row[col_map["role"]]) if "role" in col_map and col_map["role"] < len(row) else ""
            hours = parse_numeric(row[col_map["hours"]]) if "hours" in col_map and col_map["hours"] < len(row) else 0
            cust_rate = parse_numeric(row[col_map["customer_rate"]]) if "customer_rate" in col_map and col_map["customer_rate"] < len(row) else None
            kutlu_rate = parse_numeric(row[col_map["kutlu_rate"]]) if "kutlu_rate" in col_map and col_map["kutlu_rate"] < len(row) else None

            item = LineItem(
                phase=phase,
                task=task,
                role=role,
                hours=hours if hours else 0,
                customer_rate=cust_rate,
                kutlu_rate=kutlu_rate,
            )
            item.is_total = is_total
            if not is_total:
                items.append(item)
    return items


# ── Workbook generation ────────────────────────────────────────────

def apply_header(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Styles.HEADER
    c.fill = Styles.HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = Styles.THIN_BORDER
    return c


def apply_cell(ws, row, col, value, fmt=None, font=None, is_input=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or Styles.CELL
    c.border = Styles.THIN_BORDER
    if fmt:
        c.number_format = fmt
    if is_input:
        c.font = Styles.INPUT
        c.fill = Styles.INPUT_FILL
    return c


def build_rate_card(wb, items):
    """Build the Rate Card sheet. Returns {role: row_number} for VLOOKUP references."""
    ws = wb.create_sheet("Rate Card")

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"].value = "Rate Card — Role / Rate Assumptions"
    ws["A1"].font = Styles.TITLE

    ws.merge_cells("A2:D2")
    ws["A2"].value = "Blue cells are editable inputs. Change rates here to update all sheets."
    ws["A2"].font = Styles.NOTE

    # Headers at row 4
    headers = ["Role", "Customer Rate ($/hr)", "Kutlu Cost ($/hr)", "Notes"]
    for ci, h in enumerate(headers, 1):
        apply_header(ws, 4, ci, h)

    # Collect unique roles (exclude fixed-total items like Travel)
    roles = {}
    for item in items:
        if hasattr(item, '_fixed_customer_total'):
            continue
        r = item.role or "General"
        if r not in roles:
            roles[r] = {
                "customer_rate": item.effective_customer_rate(),
                "kutlu_rate": item.effective_kutlu_rate(),
            }

    # If no roles extracted, add default
    if not roles:
        roles["General"] = {
            "customer_rate": DEFAULT_CUSTOMER_RATE,
            "kutlu_rate": DEFAULT_KUTLU_RATE,
        }

    role_rows = {}
    row = 5
    for role_name, rates in roles.items():
        apply_cell(ws, row, 1, role_name)
        apply_cell(ws, row, 2, rates["customer_rate"], Styles.RATE_FMT, is_input=True)
        apply_cell(ws, row, 3, rates["kutlu_rate"], Styles.RATE_FMT, is_input=True)
        apply_cell(ws, row, 4, "Default" if rates["customer_rate"] == DEFAULT_CUSTOMER_RATE else "Custom")
        role_rows[role_name] = row
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 15

    return role_rows


def build_detail(wb, items, role_rows):
    """Build the Detail sheet with full line-item breakdown.
    Returns (first_data_row, last_data_row) for Summary formulas."""
    ws = wb.create_sheet("Detail")

    # Headers: Phase | Task | Role | Hours | Customer Rate | Customer Total | Kutlu Rate | Kutlu Total
    detail_headers = [
        "Phase", "Task/Deliverable", "Role", "Hours",
        "Customer Rate ($/hr)", "Customer Total",
        "Kutlu Rate ($/hr)", "Kutlu Total",
    ]
    for ci, h in enumerate(detail_headers, 1):
        apply_header(ws, 1, ci, h)

    rate_card_range_start = 5
    rate_card_range_end = rate_card_range_start + len(role_rows) - 1

    row = 2
    first_data = row
    for item in items:
        role_name = item.role or "General"

        apply_cell(ws, row, 1, item.phase)
        apply_cell(ws, row, 2, item.task)
        apply_cell(ws, row, 3, role_name)
        apply_cell(ws, row, 4, item.hours, Styles.HOURS_FMT, is_input=True)

        # Check for fixed-total items (e.g., travel pass-through)
        if hasattr(item, '_fixed_customer_total'):
            apply_cell(ws, row, 5, 0, Styles.RATE_FMT)
            apply_cell(ws, row, 6, item._fixed_customer_total, Styles.CURRENCY_FMT)
            apply_cell(ws, row, 7, 0, Styles.RATE_FMT)
            apply_cell(ws, row, 8, item._fixed_kutlu_total, Styles.CURRENCY_FMT)
        else:
            # Customer rate: VLOOKUP from Rate Card by role, fallback to direct input
            cust_rate_formula = (
                f"=IFERROR(VLOOKUP(C{row},'Rate Card'!A{rate_card_range_start}:B{rate_card_range_end},2,FALSE),"
                f"{item.effective_customer_rate()})"
            )
            apply_cell(ws, row, 5, cust_rate_formula, Styles.RATE_FMT)

            # Customer total = Hours * Customer Rate
            apply_cell(ws, row, 6, f"=D{row}*E{row}", Styles.CURRENCY_FMT)

            # Kutlu rate: VLOOKUP from Rate Card
            kutlu_rate_formula = (
                f"=IFERROR(VLOOKUP(C{row},'Rate Card'!A{rate_card_range_start}:C{rate_card_range_end},3,FALSE),"
                f"{item.effective_kutlu_rate()})"
            )
            apply_cell(ws, row, 7, kutlu_rate_formula, Styles.RATE_FMT)

            # Kutlu total = Hours * Kutlu Rate
            apply_cell(ws, row, 8, f"=D{row}*G{row}", Styles.CURRENCY_FMT)

        row += 1

    last_data = row - 1

    # Totals row
    total_row = row + 1
    apply_cell(ws, total_row, 1, "TOTAL", font=Styles.BOLD_CELL)
    ws.cell(row=total_row, column=1).border = Styles.TOTAL_BORDER
    for col_idx in [4, 6, 8]:
        col_letter = get_column_letter(col_idx)
        fmt = Styles.HOURS_FMT if col_idx == 4 else Styles.CURRENCY_FMT
        c = apply_cell(ws, total_row, col_idx, f"=SUM({col_letter}{first_data}:{col_letter}{last_data})", fmt, font=Styles.BOLD_CELL)
        c.border = Styles.TOTAL_BORDER
    # Fill borders on empty total cells
    for col_idx in [2, 3, 5, 7]:
        ws.cell(row=total_row, column=col_idx).border = Styles.TOTAL_BORDER

    # Column widths
    widths = [18, 35, 20, 10, 22, 18, 22, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return first_data, last_data, total_row


def build_summary(wb, items, detail_first, detail_last):
    """Build the Summary sheet — totals by phase/category."""
    ws = wb.create_sheet("Summary")

    # Collect unique phases in order
    phases = []
    seen = set()
    for item in items:
        p = item.phase or "General"
        if p not in seen:
            phases.append(p)
            seen.add(p)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Cost Proposal Summary"
    ws["A1"].font = Styles.TITLE

    # Headers at row 3
    summary_headers = [
        "Phase / Category", "Total Hours",
        "Customer Total", "Kutlu Total",
    ]
    for ci, h in enumerate(summary_headers, 1):
        apply_header(ws, 3, ci, h)

    row = 4
    first_summary = row
    for phase in phases:
        apply_cell(ws, row, 1, phase)
        # SUMIFS from Detail sheet
        apply_cell(ws, row, 2,
                   f'=SUMIFS(Detail!D{detail_first}:D{detail_last},Detail!A{detail_first}:A{detail_last},A{row})',
                   Styles.HOURS_FMT)
        apply_cell(ws, row, 3,
                   f'=SUMIFS(Detail!F{detail_first}:F{detail_last},Detail!A{detail_first}:A{detail_last},A{row})',
                   Styles.CURRENCY_FMT)
        apply_cell(ws, row, 4,
                   f'=SUMIFS(Detail!H{detail_first}:H{detail_last},Detail!A{detail_first}:A{detail_last},A{row})',
                   Styles.CURRENCY_FMT)
        row += 1

    last_summary = row - 1

    # Grand total
    total_row = row + 1
    apply_cell(ws, total_row, 1, "GRAND TOTAL", font=Styles.BOLD_CELL)
    ws.cell(row=total_row, column=1).border = Styles.TOTAL_BORDER
    for ci, col_letter in [(2, "B"), (3, "C"), (4, "D")]:
        fmt = Styles.HOURS_FMT if ci == 2 else Styles.CURRENCY_FMT
        c = apply_cell(ws, total_row, ci,
                       f"=SUM({col_letter}{first_summary}:{col_letter}{last_summary})",
                       fmt, font=Styles.BOLD_CELL)
        c.border = Styles.TOTAL_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    return total_row


def build_workbook(items, meta, output_path):
    """Assemble the complete 3-sheet workbook."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Enable full recalculation on load
    wb.calculation = CalcProperties(calcId=150000, fullCalcOnLoad=True)

    # Build sheets in order
    role_rows = build_rate_card(wb, items)
    detail_first, detail_last, detail_total = build_detail(wb, items, role_rows)
    build_summary(wb, items, detail_first, detail_last)

    # Set Summary as the active/first sheet
    wb.move_sheet("Summary", offset=-2)
    wb.active = wb.sheetnames.index("Summary")

    # Add metadata as print header if available
    proposal_title = meta.get("Proposal.Title", meta.get("RFP.Title", meta.get("title", "")))
    vendor = meta.get("Vendor.Name", meta.get("client", {}).get("name", "") if isinstance(meta.get("client"), dict) else "")
    if proposal_title:
        for ws in wb.worksheets:
            ws.oddHeader.center.text = proposal_title
            if vendor:
                ws.oddHeader.right.text = vendor

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f"OK: {output_path}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Line items: {len(items)}")
    print(f"  Roles: {len(role_rows)}")
    return True


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Build cost proposal workbook from markdown tables."
    )
    parser.add_argument("markdown_file", help="Markdown file with cost tables")
    parser.add_argument("output_xlsx", nargs="?", help="Output .xlsx path (auto-derived if omitted)")
    args = parser.parse_args()

    md_path = os.path.abspath(args.markdown_file)
    if not os.path.isfile(md_path):
        print(f"ERROR: File not found: {md_path}")
        sys.exit(1)

    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    meta, body = parse_frontmatter(text)
    flat = flatten_dict(meta)

    # Strategy 1: Extract from YAML pricing block
    items = extract_from_yaml_pricing(meta)
    if items:
        print(f"  Source: YAML pricing block ({len(items)} phases)")
    else:
        # Strategy 2: Extract from <!-- cost-start/end --> markers or markdown tables
        cost_md = extract_cost_markdown(body)
        if not cost_md:
            cost_md = body

        tables = parse_md_tables(cost_md)
        if not tables:
            print("ERROR: No pricing data found. Expected either:")
            print("  - YAML frontmatter with 'pricing.phases' block")
            print("  - Markdown tables (with optional <!-- cost-start/end --> markers)")
            sys.exit(1)

        items = extract_line_items(tables)
        if not items:
            print("ERROR: No line items extracted from tables.")
            sys.exit(1)
        print(f"  Source: markdown tables ({len(items)} items)")

    # Derive output path
    if args.output_xlsx:
        out_path = os.path.abspath(args.output_xlsx)
    else:
        md_dir = os.path.dirname(md_path)
        rfp_number = flat.get("RFP.Number", flat.get("opportunity_id", "")).strip()
        title = flat.get("Proposal.Title", flat.get("RFP.Title", flat.get("title", "Cost Proposal"))).strip()
        if rfp_number:
            name = sanitize_filename(f"Cost Proposal - {rfp_number} {title}.xlsx")
        else:
            name = sanitize_filename(f"Cost Proposal - {title}.xlsx")
        out_path = os.path.join(md_dir, name)

    build_workbook(items, flat, out_path)


if __name__ == "__main__":
    main()
